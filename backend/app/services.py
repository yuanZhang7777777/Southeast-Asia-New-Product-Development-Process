from __future__ import annotations

from io import BytesIO
from collections import defaultdict
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app import models, schemas


def audit(db: Session, action: str, entity_type: str, entity_id: str | None, detail: dict, actor_name: str | None = None) -> None:
    db.add(models.AuditLog(action=action, entity_type=entity_type, entity_id=entity_id, detail=detail, actor_name=actor_name))


def create_opportunity(db: Session, payload: schemas.OpportunityCreate) -> models.NewProductOpportunity:
    item = models.NewProductOpportunity(**payload.model_dump())
    db.add(item)
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        query = select(models.NewProductOpportunity).where(
            models.NewProductOpportunity.source_type == payload.source_type,
            models.NewProductOpportunity.source_file == payload.source_file,
            models.NewProductOpportunity.source_sheet == payload.source_sheet,
            models.NewProductOpportunity.source_row == payload.source_row,
        )
        existing = db.scalar(query)
        if existing is None:
            raise
        return existing
    db.add(
        models.SourceRecordSnapshot(
            opportunity_id=item.id,
            source_file=item.source_file,
            source_sheet=item.source_sheet,
            source_row=item.source_row,
            column_range="source_row",
            payload=payload.snapshot or payload.model_dump(),
        )
    )
    audit(db, "opportunity.created", "new_product_opportunity", item.id, {"main_sku": item.main_sku, "sub_sku": item.sub_sku})
    return item


def preview_assignments(
    opportunities: list[models.NewProductOpportunity], candidates: list[str]
) -> list[schemas.AssignmentPreviewItem]:
    if not candidates:
        return []
    grouped: dict[str, list[models.NewProductOpportunity]] = defaultdict(list)
    for item in opportunities:
        grouped[item.main_sku].append(item)

    loads = {candidate: 0 for candidate in candidates}
    output: list[schemas.AssignmentPreviewItem] = []
    for main_sku, items in sorted(grouped.items(), key=lambda pair: len(pair[1]), reverse=True):
        assignee = min(loads, key=lambda name: (loads[name], name))
        count = len(items)
        loads[assignee] += count
        output.append(schemas.AssignmentPreviewItem(main_sku=main_sku, sub_sku_count=count, suggested_assignee=assignee))
    return output


def confirm_assignment(db: Session, payload: schemas.AssignmentConfirmRequest) -> list[models.FlowTask]:
    opportunities = list(
        db.scalars(select(models.NewProductOpportunity).where(models.NewProductOpportunity.id.in_(payload.opportunity_ids)))
    )
    tasks: list[models.FlowTask] = []
    for opportunity in opportunities:
        flow = models.FlowInstance(
            opportunity_id=opportunity.id,
            current_node="sales_claim",
            current_status="assigned",
            owner_user_id=payload.assignee_user_id,
            owner_role="sales",
            deadline_at=payload.deadline_at,
        )
        db.add(flow)
        db.flush()
        task = models.FlowTask(
            flow_instance_id=flow.id,
            node_code="sales_claim",
            task_type="sales_claim",
            assignee_user_id=payload.assignee_user_id,
            assignee_name=payload.assignee_name,
            assignee_role="sales",
            deadline_at=payload.deadline_at,
        )
        opportunity.current_status = "assigned"
        db.add(task)
        tasks.append(task)
        audit(db, "assignment.confirmed", "flow_task", task.id, {"opportunity_id": opportunity.id}, payload.assignee_name)
    return tasks


def submit_claim(db: Session, payload: schemas.ClaimCreate) -> models.SalesClaimForecast:
    if payload.claim_result == "claim" and payload.claim_daily_sales is None:
        raise ValueError("claim_daily_sales is required when claim_result is claim")
    if payload.claim_result == "reject" and not payload.reject_reason:
        raise ValueError("reject_reason is required when claim_result is reject")

    claim = models.SalesClaimForecast(
        opportunity_id=payload.opportunity_id,
        platform="Shopee",
        salesperson_name=payload.salesperson_name,
        claim_result=payload.claim_result,
        claim_daily_sales=payload.claim_daily_sales,
        reject_reason=payload.reject_reason,
        feedback_summary=payload.feedback_summary,
        source_column="platform",
    )
    db.add(claim)
    opportunity = db.get(models.NewProductOpportunity, payload.opportunity_id)
    if opportunity:
        opportunity.current_status = "claim_submitted" if payload.claim_result == "claim" else "claim_rejected"
    task = db.scalar(
        select(models.FlowTask)
        .join(models.FlowInstance)
        .where(
            models.FlowInstance.opportunity_id == payload.opportunity_id,
            models.FlowTask.task_type == "sales_claim",
            models.FlowTask.status == "pending",
        )
        .order_by(models.FlowTask.created_at.desc())
    )
    if task:
        task.status = "completed"
        task.completed_at = datetime.now(timezone.utc)
    create_review_task(db, payload.opportunity_id, payload.salesperson_name)
    audit(db, "claim.submitted", "sales_claim_forecast", claim.id, payload.model_dump(), payload.salesperson_name)
    return claim


def create_review_task(db: Session, opportunity_id: str, actor_name: str | None = None) -> None:
    flow = db.scalar(
        select(models.FlowInstance)
        .where(models.FlowInstance.opportunity_id == opportunity_id)
        .order_by(models.FlowInstance.created_at.desc())
    )
    if not flow:
        flow = models.FlowInstance(opportunity_id=opportunity_id, current_node="review", current_status="review_pending")
        db.add(flow)
        db.flush()
    flow.current_node = "review"
    flow.current_status = "review_pending"
    task = models.FlowTask(
        flow_instance_id=flow.id,
        node_code="review",
        task_type="manager_review",
        assignee_role="manager",
    )
    db.add(task)
    audit(db, "review_task.created", "flow_task", task.id, {"opportunity_id": opportunity_id}, actor_name)


def submit_review(db: Session, payload: schemas.ReviewCreate) -> models.ReviewRecord:
    record = models.ReviewRecord(
        opportunity_id=payload.opportunity_id,
        reviewer_name=payload.reviewer_name,
        review_status=payload.review_status,
        review_comment=payload.review_comment,
    )
    db.add(record)
    opportunity = db.get(models.NewProductOpportunity, payload.opportunity_id)
    if opportunity:
        opportunity.current_status = "ready_for_stocking" if payload.review_status == "approved" else f"review_{payload.review_status}"
    task = db.scalar(
        select(models.FlowTask)
        .join(models.FlowInstance)
        .where(
            models.FlowInstance.opportunity_id == payload.opportunity_id,
            models.FlowTask.task_type == "manager_review",
            models.FlowTask.status == "pending",
        )
        .order_by(models.FlowTask.created_at.desc())
    )
    if task:
        task.status = "completed"
        task.completed_at = datetime.now(timezone.utc)
    if payload.review_status == "approved":
        audit(db, "opportunity.ready_for_stocking", "new_product_opportunity", payload.opportunity_id, {}, payload.reviewer_name)
    audit(db, "review.submitted", "review_record", record.id, payload.model_dump(), payload.reviewer_name)
    return record


def create_stocking_draft_from_claim(db: Session, opportunity_id: str, actor_name: str | None = None) -> models.StockingRequest | None:
    claim = db.scalar(
        select(models.SalesClaimForecast)
        .where(
            models.SalesClaimForecast.opportunity_id == opportunity_id,
            models.SalesClaimForecast.claim_result == "claim",
        )
        .order_by(models.SalesClaimForecast.created_at.desc())
    )
    opportunity = db.get(models.NewProductOpportunity, opportunity_id)
    if not claim or not claim.claim_daily_sales or not opportunity:
        return None
    quantity = int(round(claim.claim_daily_sales * 30))
    request = models.StockingRequest(
        opportunity_id=opportunity_id,
        salesperson_name=claim.salesperson_name,
        main_sku=opportunity.main_sku,
        sub_sku=opportunity.sub_sku,
        daily_sales=claim.claim_daily_sales,
        quantity=quantity,
        country=opportunity.country,
        status="draft",
    )
    db.add(request)
    audit(db, "stocking.draft_created", "stocking_request", request.id, {"quantity": quantity}, actor_name)
    return request


def list_available_stocking_items(db: Session) -> list[schemas.AvailableStockingItem]:
    opportunities = list(
        db.scalars(
            select(models.NewProductOpportunity)
            .where(models.NewProductOpportunity.current_status == "ready_for_stocking")
            .order_by(models.NewProductOpportunity.updated_at.desc())
        )
    )
    items: list[schemas.AvailableStockingItem] = []
    for opportunity in opportunities:
        claim = latest_claim(db, opportunity.id)
        if not claim or claim.claim_daily_sales is None:
            continue
        review = latest_approved_review(db, opportunity.id)
        quantity = int(round(claim.claim_daily_sales * 30))
        items.append(
            schemas.AvailableStockingItem(
                time=review.created_at if review else datetime.now(timezone.utc),
                selection_source=selection_source_label(opportunity),
                salesperson_name=claim.salesperson_name or claim_prefill_salesperson(opportunity),
                main_sku=opportunity.main_sku,
                sub_sku=opportunity.sub_sku,
                site=opportunity.site,
                claim_daily_sales=claim.claim_daily_sales,
                quantity=quantity,
            )
        )
    return items


def build_available_stocking_workbook(items: list[schemas.AvailableStockingItem]) -> bytes:
    headers = [
        "操作状态",
        "时间",
        "备货类型",
        "选品数据源",
        "销售员",
        "主SKU",
        "子SKU",
        "站点",
        "认领单销",
        "备货量",
        "备货国家",
        "仓库",
        "成本价",
        "单个体积",
        "货值",
        "补货原因",
        "★是否需要开品邮件",
        "★开品邮件状态",
    ]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "可备货清单"
    worksheet.append(headers)
    for item in items:
        worksheet.append(
            [
                item.operation_status,
                item.time.replace(tzinfo=None),
                item.stocking_type,
                item.selection_source,
                item.salesperson_name,
                item.main_sku,
                item.sub_sku,
                item.site,
                item.claim_daily_sales,
                item.quantity,
                item.stocking_country,
                item.warehouse,
                item.cost_price,
                item.unit_volume,
                item.amount,
                item.replenishment_reason,
                item.needs_launch_email,
                item.launch_email_status,
            ]
        )

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 32)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    for cell in worksheet["B"][1:]:
        cell.number_format = "yyyy-mm-dd hh:mm"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def latest_claim(db: Session, opportunity_id: str) -> models.SalesClaimForecast | None:
    return db.scalar(
        select(models.SalesClaimForecast)
        .where(
            models.SalesClaimForecast.opportunity_id == opportunity_id,
            models.SalesClaimForecast.claim_result == "claim",
            models.SalesClaimForecast.source_column == "platform",
        )
        .order_by(models.SalesClaimForecast.created_at.desc())
    )


def latest_approved_review(db: Session, opportunity_id: str) -> models.ReviewRecord | None:
    return db.scalar(
        select(models.ReviewRecord)
        .where(
            models.ReviewRecord.opportunity_id == opportunity_id,
            models.ReviewRecord.review_status == "approved",
        )
        .order_by(models.ReviewRecord.created_at.desc())
    )


def selection_source_label(opportunity: models.NewProductOpportunity) -> str:
    sheet = opportunity.source_sheet
    if opportunity.source_type == "selection1_developer_claim_feedback":
        label = "选品1-开发部门认领反馈"
    else:
        label = opportunity.source_type
    return f"{label} + {sheet}" if sheet else label


def claim_prefill_salesperson(opportunity: models.NewProductOpportunity) -> str | None:
    snapshot = opportunity.snapshot or {}
    claim = snapshot.get("claim_prefill") if isinstance(snapshot, dict) else None
    if isinstance(claim, dict):
        value = claim.get("salesperson_name")
        return str(value) if value else None
    cells = snapshot.get("cells") if isinstance(snapshot, dict) else None
    if isinstance(cells, dict):
        value = cells.get("CD")
        return str(value) if value else None
    return None


def create_notification_log(db: Session, payload: schemas.NotificationTestRequest) -> models.NotificationLog:
    existing = db.scalar(select(models.NotificationLog).where(models.NotificationLog.dedupe_key == payload.dedupe_key))
    if existing:
        return existing
    item = models.NotificationLog(
        dedupe_key=payload.dedupe_key,
        receiver_name=payload.receiver_name,
        channel=payload.channel,
        message_title=payload.title,
        send_status="logged",
    )
    db.add(item)
    audit(db, "notification.logged", "notification_log", item.id, payload.model_dump(), payload.receiver_name)
    return item


def reassign_task(db: Session, payload: schemas.ReassignRequest) -> models.FlowTask:
    task = db.get(models.FlowTask, payload.task_id)
    if task is None:
        raise LookupError("task not found")
    old_assignee = task.assignee_name
    task.assignee_name = payload.assignee_name
    task.assignee_user_id = payload.assignee_user_id
    task.status = "pending"
    audit(
        db,
        "task.reassigned",
        "flow_task",
        task.id,
        {"old_assignee": old_assignee, "new_assignee": payload.assignee_name, "reason": payload.reason},
        payload.assignee_name,
    )
    return task


def due_summary_date(base: datetime | None = None) -> datetime:
    start = base or datetime.now(timezone.utc)
    return start.replace(microsecond=0)
