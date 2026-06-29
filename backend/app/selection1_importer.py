from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app import models, schemas, services


SOURCE_TYPE = "selection1_developer_claim_feedback"
SOURCE_LABEL = "选品1-开发部门认领反馈"
DEFAULT_SHEET = "开发0623期"

MAIN_COLUMNS = {
    "A": "site",
    "B": "developer_department",
    "C": "developer_name",
    "D": "category_level1",
    "E": "keyword",
    "F": "image_url",
    "G": "main_sku_name",
    "H": "main_sku",
    "I": "sub_sku_name",
    "J": "sub_sku",
    "K": "product_type",
    "L": "reason",
}

MARKET_GROUPS = [
    ("最低价", "Z", "AA", "AB"),
    ("most_orders", "AC", "AD", "AE"),
    ("月销次高", "AF", "AG", "AH"),
    ("月销第三高", "AI", "AJ", "AK"),
    ("新晋", "AL", "AM", "AN"),
]

PRICING_SNAPSHOT_COLUMNS = {
    "AO": "参考单销",
    "AP": "参考定价",
    "AQ": "一次毛利额THB",
    "AR": "一次毛利额RMB",
    "AS": "一次毛利率",
    "AT": "预估单销",
    "AU": "推广期定价",
    "AV": "推广期利润率",
}

CLAIM_COLUMNS = {
    "CC": "reject_reason",
    "CD": "salesperson_name",
    "CE": "claim_result",
    "CF": "claim_daily_sales",
    "CG": "feedback_summary",
    "CH": "note",
}

SNAPSHOT_COLUMNS = tuple(
    list(MAIN_COLUMNS)
    + [column for group in MARKET_GROUPS for column in group[1:]]
    + list(PRICING_SNAPSHOT_COLUMNS)
    + list(CLAIM_COLUMNS)
)
MAX_SOURCE_COLUMN = column_index_from_string("CH")

COUNTRY_BY_SITE = {
    "菲律宾": "PH",
    "菲": "PH",
    "PH": "PH",
    "泰国": "TH",
    "泰": "TH",
    "TH": "TH",
    "越南": "VN",
    "越": "VN",
    "VN": "VN",
    "马来西亚": "MY",
    "马来": "MY",
    "MY": "MY",
}


def import_selection1_workbook(db: Session, payload: schemas.Selection1ImportRequest) -> schemas.Selection1ImportResponse:
    source_path = resolve_source_file(payload.source_file)
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    if payload.source_sheet not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"sheet not found: {payload.source_sheet}; available sheets: {available}")

    worksheet = workbook[payload.source_sheet]
    try:
        worksheet.reset_dimensions()
    except AttributeError:
        pass

    created_count = 0
    updated_count = 0
    skipped_count = 0
    market_research_count = 0
    prefill_claim_count = 0
    task_count = 0
    processed_count = 0

    for source_row, row in enumerate(worksheet.iter_rows(min_row=3, max_col=MAX_SOURCE_COLUMN, values_only=True), start=3):
        if payload.max_rows is not None and processed_count >= payload.max_rows:
            break
        parsed = parse_selection1_row(row)
        if parsed is None:
            skipped_count += 1
            continue
        processed_count += 1

        opportunity, created = upsert_opportunity(db, parsed, source_path.name, payload.source_sheet, source_row)
        created_count += int(created)
        updated_count += int(not created)

        market_research_count += replace_market_research(db, opportunity.id, parsed)
        if replace_source_claim_prefill(db, opportunity.id, parsed):
            prefill_claim_count += 1
        if ensure_import_claim_task(db, opportunity, parsed):
            task_count += 1

    services.audit(
        db,
        "selection1.imported",
        "new_product_opportunity",
        None,
        {
            "source_file": source_path.name,
            "source_sheet": payload.source_sheet,
            "imported_count": created_count + updated_count,
            "created_count": created_count,
            "updated_count": updated_count,
            "skipped_count": skipped_count,
        },
    )

    return schemas.Selection1ImportResponse(
        source_file=source_path.name,
        source_sheet=payload.source_sheet,
        imported_count=created_count + updated_count,
        created_count=created_count,
        updated_count=updated_count,
        skipped_count=skipped_count,
        market_research_count=market_research_count,
        prefill_claim_count=prefill_claim_count,
        task_count=task_count,
    )


def resolve_source_file(source_file: str | None) -> Path:
    if source_file:
        path = Path(source_file)
        if not path.is_absolute():
            path = Path.cwd() / path
        if not path.exists():
            raise FileNotFoundError(str(path))
        return path

    search_roots = [Path.cwd(), Path.cwd().parent]
    for root in search_roots:
        matches = sorted(root.glob("选品1：海外仓开发部门开发新品认领-反馈*.xlsx"))
        if matches:
            return matches[0]
    raise FileNotFoundError("selection1 source workbook not found")


def parse_selection1_row(row: tuple[Any, ...]) -> dict[str, Any] | None:
    values = {column: clean_cell(cell_value(row, column)) for column in SNAPSHOT_COLUMNS}
    main_sku = text_value(values["H"])
    sub_sku = text_value(values["J"])
    if not main_sku or not sub_sku:
        return None
    if is_summary_row(values):
        return None

    site = text_value(values["A"])
    pricing_snapshot = {label: values[column] for column, label in PRICING_SNAPSHOT_COLUMNS.items() if values[column] not in (None, "")}
    feedback_parts = [text_value(values["CG"]), text_value(values["CH"])]
    claim_prefill = {
        "reject_reason": text_value(values["CC"]),
        "salesperson_name": text_value(values["CD"]),
        "claim_result": normalize_claim_result(values["CE"]),
        "claim_daily_sales": number_value(values["CF"]),
        "feedback_summary": "\n".join(part for part in feedback_parts if part),
    }
    parsed = {
        "main": {field: values[column] for column, field in MAIN_COLUMNS.items()},
        "country": derive_country(site),
        "market_items": parse_market_items(values),
        "reference_daily_sales": number_value(values["AO"]),
        "reference_price": number_value(values["AP"]),
        "pricing_snapshot": pricing_snapshot,
        "claim_prefill": claim_prefill,
        "snapshot": {
            "source_type": SOURCE_TYPE,
            "allowed_columns": list(SNAPSHOT_COLUMNS),
            "cells": values,
            "pricing_snapshot": pricing_snapshot,
            "claim_prefill": claim_prefill,
        },
    }
    return parsed


def upsert_opportunity(
    db: Session, parsed: dict[str, Any], source_file: str, source_sheet: str, source_row: int
) -> tuple[models.NewProductOpportunity, bool]:
    existing = db.scalar(
        select(models.NewProductOpportunity).where(
            models.NewProductOpportunity.source_type == SOURCE_TYPE,
            models.NewProductOpportunity.source_file == source_file,
            models.NewProductOpportunity.source_sheet == source_sheet,
            models.NewProductOpportunity.source_row == source_row,
        )
    )
    opportunity_data = {
        **parsed["main"],
        "source_type": SOURCE_TYPE,
        "source_file": source_file,
        "source_sheet": source_sheet,
        "source_row": source_row,
        "batch": source_sheet,
        "country": parsed["country"],
        "snapshot": parsed["snapshot"],
    }
    if existing:
        for field, value in opportunity_data.items():
            if field != "current_status":
                setattr(existing, field, value)
        snapshot = db.scalar(
            select(models.SourceRecordSnapshot).where(
                models.SourceRecordSnapshot.opportunity_id == existing.id,
                models.SourceRecordSnapshot.column_range == "A:L,Z:AN,AO:AV,CC:CH",
            )
        )
        if snapshot:
            snapshot.payload = parsed["snapshot"]
        else:
            add_source_snapshot(db, existing)
        return existing, False

    opportunity = models.NewProductOpportunity(**opportunity_data)
    db.add(opportunity)
    db.flush()
    add_source_snapshot(db, opportunity)
    services.audit(
        db,
        "opportunity.created",
        "new_product_opportunity",
        opportunity.id,
        {"main_sku": opportunity.main_sku, "sub_sku": opportunity.sub_sku, "source_type": SOURCE_TYPE},
    )
    return opportunity, True


def add_source_snapshot(db: Session, opportunity: models.NewProductOpportunity) -> None:
    db.add(
        models.SourceRecordSnapshot(
            opportunity_id=opportunity.id,
            source_file=opportunity.source_file,
            source_sheet=opportunity.source_sheet,
            source_row=opportunity.source_row,
            column_range="A:L,Z:AN,AO:AV,CC:CH",
            payload=opportunity.snapshot,
        )
    )


def replace_market_research(db: Session, opportunity_id: str, parsed: dict[str, Any]) -> int:
    db.execute(delete(models.MarketResearchItem).where(models.MarketResearchItem.opportunity_id == opportunity_id))
    count = 0
    for item in parsed["market_items"]:
        db.add(
            models.MarketResearchItem(
                opportunity_id=opportunity_id,
                research_type=item["research_type"],
                competitor_url=item["competitor_url"],
                competitor_price=item["competitor_price"],
                competitor_monthly_sales=item["competitor_monthly_sales"],
                reference_daily_sales=parsed["reference_daily_sales"],
                reference_price=parsed["reference_price"],
            )
        )
        count += 1
    return count


def replace_source_claim_prefill(db: Session, opportunity_id: str, parsed: dict[str, Any]) -> bool:
    claim = parsed["claim_prefill"]
    db.execute(
        delete(models.SalesClaimForecast).where(
            models.SalesClaimForecast.opportunity_id == opportunity_id,
            models.SalesClaimForecast.source_column == "CC:CH",
        )
    )
    if not any(claim.values()):
        return False
    db.add(
        models.SalesClaimForecast(
            opportunity_id=opportunity_id,
            platform="Shopee",
            salesperson_name=claim["salesperson_name"],
            claim_result=claim["claim_result"],
            claim_daily_sales=claim["claim_daily_sales"],
            reject_reason=claim["reject_reason"],
            feedback_summary=claim["feedback_summary"],
            source_column="CC:CH",
        )
    )
    return True


def ensure_import_claim_task(db: Session, opportunity: models.NewProductOpportunity, parsed: dict[str, Any]) -> bool:
    task = db.scalar(
        select(models.FlowTask)
        .join(models.FlowInstance)
        .where(
            models.FlowInstance.opportunity_id == opportunity.id,
            models.FlowTask.task_type == "sales_claim",
        )
        .order_by(models.FlowTask.created_at.desc())
    )
    assignee_name = parsed["claim_prefill"]["salesperson_name"]
    status = "assigned" if assignee_name else "open_claim_pool"
    if task:
        if task.status == "pending" and assignee_name and not task.assignee_name:
            task.assignee_name = assignee_name
        if opportunity.current_status in {"pending_assignment", "open_claim_pool", "assigned"}:
            opportunity.current_status = status
        return False

    flow = models.FlowInstance(
        opportunity_id=opportunity.id,
        current_node="sales_claim",
        current_status=status,
        owner_role="sales",
    )
    db.add(flow)
    db.flush()
    db.add(
        models.FlowTask(
            flow_instance_id=flow.id,
            node_code="sales_claim",
            task_type="sales_claim",
            assignee_name=assignee_name,
            assignee_role="sales",
        )
    )
    if opportunity.current_status in {"pending_assignment", "open_claim_pool", "assigned"}:
        opportunity.current_status = status
    return True


def parse_market_items(values: dict[str, Any]) -> list[dict[str, Any]]:
    items = []
    for research_type, url_column, price_column, sales_column in MARKET_GROUPS:
        competitor_url = text_value(values[url_column])
        competitor_price = number_value(values[price_column])
        if not competitor_url and competitor_price is None:
            continue
        items.append(
            {
                "research_type": research_type,
                "competitor_url": competitor_url,
                "competitor_price": competitor_price,
                "competitor_monthly_sales": number_value(values[sales_column]),
            }
        )
    return items


def is_summary_row(values: dict[str, Any]) -> bool:
    joined = " ".join(str(value) for value in values.values() if value is not None)
    return "小计" in joined or "合计" in joined


def derive_country(site: str | None) -> str | None:
    if not site:
        return None
    cleaned = site.strip().upper()
    return COUNTRY_BY_SITE.get(site.strip()) or COUNTRY_BY_SITE.get(cleaned) or cleaned


def normalize_claim_result(value: Any) -> str | None:
    text = text_value(value)
    if text in {"是", "认领", "claim", "CLAIM", "yes", "YES"}:
        return "claim"
    if text in {"否", "不认领", "reject", "REJECT", "no", "NO"}:
        return "reject"
    return None


def clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def cell_value(row: tuple[Any, ...], column: str) -> Any:
    index = column_index_from_string(column) - 1
    if index >= len(row):
        return None
    return row[index]


def text_value(value: Any) -> str | None:
    cleaned = clean_cell(value)
    if cleaned is None:
        return None
    return str(cleaned).strip()


def number_value(value: Any) -> float | None:
    cleaned = clean_cell(value)
    if cleaned is None:
        return None
    if isinstance(cleaned, int | float):
        return float(cleaned)
    text = str(cleaned).strip().replace(",", "")
    if not text or text in {"-", "/"}:
        return None
    number_text = "".join(ch for ch in text if ch.isdigit() or ch in ".-")
    if number_text in {"", ".", "-", "-."}:
        return None
    try:
        return float(number_text)
    except ValueError:
        return None

