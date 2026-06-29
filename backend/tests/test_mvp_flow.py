import os
import sys
from pathlib import Path

os.environ["DATABASE_URL"] = f"sqlite:///{Path(__file__).with_name('test_workflow.db')}"
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402

from app.db import Base, engine  # noqa: E402
from app.main import app  # noqa: E402
from app import models  # noqa: E402
from app.db import SessionLocal  # noqa: E402


client = TestClient(app)


def setup_function() -> None:
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)


def test_health() -> None:
    response = client.get("/health")

    assert response.status_code == 200
    assert response.json()["status"] == "ok"


def test_mvp_flow_and_notification_dedupe() -> None:
    import_response = client.post(
        "/opportunities/import",
        json={
            "items": [
                {
                    "source_type": "ph_site_sheet",
                    "source_file": "东南亚海外仓新品表-PH.xlsx",
                    "source_sheet": "新品",
                    "source_row": 2,
                    "country": "PH",
                    "site": "PH",
                    "main_sku": "MAIN-001",
                    "sub_sku": "SUB-001",
                    "keyword": "storage rack",
                    "snapshot": {"A": "MAIN-001", "B": "SUB-001"},
                },
                {
                    "source_type": "ph_site_sheet",
                    "source_file": "东南亚海外仓新品表-PH.xlsx",
                    "source_sheet": "新品",
                    "source_row": 3,
                    "country": "PH",
                    "site": "PH",
                    "main_sku": "MAIN-001",
                    "sub_sku": "SUB-002",
                    "keyword": "storage rack",
                    "snapshot": {"A": "MAIN-001", "B": "SUB-002"},
                },
            ]
        },
    )
    assert import_response.status_code == 200
    opportunity_ids = [item["id"] for item in import_response.json()]

    preview_response = client.post(
        "/assignments/preview",
        json={"opportunity_ids": opportunity_ids, "candidates": ["销售A", "销售B"]},
    )
    assert preview_response.status_code == 200
    assert preview_response.json()["items"] == [
        {"main_sku": "MAIN-001", "sub_sku_count": 2, "suggested_assignee": "销售A"}
    ]

    confirm_response = client.post(
        "/assignments/confirm",
        json={"opportunity_ids": opportunity_ids, "assignee_name": "销售A"},
    )
    assert confirm_response.status_code == 200
    assert len(confirm_response.json()) == 2

    claim_response = client.post(
        "/claims",
        json={
            "opportunity_id": opportunity_ids[0],
            "salesperson_name": "销售A",
            "claim_result": "claim",
            "claim_daily_sales": 2,
            "feedback_summary": "可认领",
        },
    )
    assert claim_response.status_code == 200

    review_response = client.post(
        "/reviews",
        json={
            "opportunity_id": opportunity_ids[0],
            "reviewer_name": "主管A",
            "review_status": "approved",
            "review_comment": "通过",
        },
    )
    assert review_response.status_code == 200

    stocking_response = client.get("/stocking/requests")
    assert stocking_response.status_code == 200
    assert stocking_response.json() == []

    available_response = client.get("/stocking/available-list")
    assert available_response.status_code == 200
    available_items = available_response.json()
    assert available_items[0]["claim_daily_sales"] == 2
    assert available_items[0]["quantity"] == 60

    export_response = client.get("/stocking/available-list/export")
    assert export_response.status_code == 200
    assert export_response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    first_notice = client.post(
        "/notifications/test",
        json={"receiver_name": "销售A", "title": "测试提醒", "dedupe_key": "notice:claim:1"},
    )
    second_notice = client.post(
        "/notifications/test",
        json={"receiver_name": "销售A", "title": "测试提醒", "dedupe_key": "notice:claim:1"},
    )
    assert first_notice.status_code == 200
    assert second_notice.status_code == 200
    assert first_notice.json()["id"] == second_notice.json()["id"]
    assert len(client.get("/notifications/logs").json()) == 1


def test_selection1_import_is_idempotent_and_exportable(tmp_path: Path) -> None:
    workbook_path = tmp_path / "selection1.xlsx"
    build_selection1_fixture(workbook_path)

    payload = {"source_file": str(workbook_path), "source_sheet": "开发0623期"}
    first_import = client.post("/opportunities/import/selection1", json=payload)
    second_import = client.post("/opportunities/import/selection1", json=payload)

    assert first_import.status_code == 200
    assert second_import.status_code == 200
    assert first_import.json()["imported_count"] == 1
    assert first_import.json()["market_research_count"] == 2
    assert first_import.json()["prefill_claim_count"] == 1
    assert second_import.json()["created_count"] == 0
    assert second_import.json()["updated_count"] == 1

    opportunities = client.get("/opportunities").json()
    assert len(opportunities) == 1
    opportunity = opportunities[0]
    assert opportunity["site"] == "泰国"
    assert opportunity["country"] == "TH"
    assert opportunity["main_sku"] == "MAIN-001"
    assert opportunity["sub_sku"] == "SUB-001"
    assert opportunity["current_status"] == "assigned"

    with SessionLocal() as db:
        market_items = db.query(models.MarketResearchItem).all()
        source_claims = db.query(models.SalesClaimForecast).filter_by(source_column="CC:CH").all()
        snapshots = db.query(models.SourceRecordSnapshot).all()
    assert len(market_items) == 2
    assert len(source_claims) == 1
    assert snapshots[0].column_range == "A:L,Z:AN,AO:AV,CC:CH"
    assert "M" not in snapshots[0].payload["cells"]
    assert "AW" not in snapshots[0].payload["cells"]
    assert "CB" not in snapshots[0].payload["cells"]

    claim_response = client.post(
        "/claims",
        json={
            "opportunity_id": opportunity["id"],
            "salesperson_name": "销售A",
            "claim_result": "claim",
            "claim_daily_sales": 3,
            "feedback_summary": "平台提交",
        },
    )
    assert claim_response.status_code == 200

    review_response = client.post(
        "/reviews",
        json={
            "opportunity_id": opportunity["id"],
            "reviewer_name": "主管A",
            "review_status": "approved",
            "review_comment": "通过",
        },
    )
    assert review_response.status_code == 200
    assert client.get("/stocking/requests").json() == []

    export_path = tmp_path / "available.xlsx"
    export_response = client.get("/stocking/available-list/export")
    export_path.write_bytes(export_response.content)
    exported = load_workbook(export_path, data_only=True)
    sheet = exported["可备货清单"]
    headers = [cell.value for cell in sheet[1]]
    values = [cell.value for cell in sheet[2]]
    row = dict(zip(headers, values))
    assert row["销售员"] == "销售A"
    assert row["主SKU"] == "MAIN-001"
    assert row["子SKU"] == "SUB-001"
    assert row["认领单销"] == 3
    assert row["备货量"] == 90
    assert row["成本价"] is None
    assert row["单个体积"] is None
    assert row["货值"] is None
    assert "★是否需要开品邮件" in headers
    assert "★开品邮件状态" in headers


def build_selection1_fixture(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "开发0623期"
    worksheet.append([None] * 86)
    headers = [None] * 86
    for column_index, title in {
        1: "站点",
        2: "部门",
        3: "开发员",
        4: "一级类目",
        5: "关键词",
        6: "产品图片",
        7: "主SKU名称",
        8: "主SKU",
        9: "子SKU名称",
        10: "子SKU",
        11: "产品类型",
        12: "开品理由",
        26: "最低价链接",
        27: "售价1",
        28: "月销1",
        29: "most orders链接",
        30: "售价2",
        31: "月销2",
        41: "参考单销",
        42: "参考定价(THB)",
        43: "一次毛利额(THB)",
        81: "不认领理由",
        82: "主销售员",
        83: "是否认领",
        84: "认领单销",
        85: "销售反馈总结",
        86: "备注",
    }.items():
        headers[column_index - 1] = title
    worksheet.append(headers)
    row = [None] * 86
    for column_index, value in {
        1: "泰国",
        2: "产品开发六部",
        3: "开发A",
        4: "汽配与摩配",
        5: "Key Lock Cap Cover",
        6: "https://example.com/image.jpg",
        7: "摩托车改装电门锁保护盖",
        8: "MAIN-001",
        9: "红色",
        10: "SUB-001",
        11: "利润",
        12: "开品理由",
        26: "https://example.com/low",
        27: 105,
        28: 1840,
        29: "https://example.com/orders",
        30: 119,
        31: 5904,
        41: 2,
        42: 105,
        43: 10.5,
        81: "",
        82: "销售A",
        83: "是",
        84: 1,
        85: "源表反馈",
        86: "源表备注",
    }.items():
        row[column_index - 1] = value
    worksheet.append(row)
    worksheet.append([None] * 86)
    workbook.save(path)
